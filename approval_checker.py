import imaplib
from add_stock_macro import add_stock_via_vba
import email
import re
import openpyxl
from datetime import datetime
import yfinance as yf
from config import EMAIL_SENDER, EMAIL_PASSWORD, EXCEL_FILE_PATH, HOLDINGS_SHEET, NSE_SUFFIX
from email_handler import send_report_email


def fetch_approval_replies():
    approvals_with_qty = []
    approvals_need_qty = []
    rejections = []

    try:
        mail = imaplib.IMAP4_SSL('imap.gmail.com')
        mail.login(EMAIL_SENDER, EMAIL_PASSWORD)
        mail.select('inbox')
        _, messages = mail.search(None, 'UNSEEN')

        for msg_id in messages[0].split():
            _, msg_data = mail.fetch(msg_id, '(RFC822)')
            msg = email.message_from_bytes(msg_data[0][1])

            body = ''
            if msg.is_multipart():
                for part in msg.walk():
                    if part.get_content_type() == 'text/plain':
                        body = part.get_payload(decode=True).decode('utf-8', errors='ignore')
                        break
            else:
                body = msg.get_payload(decode=True).decode('utf-8', errors='ignore')

            first_line = body.strip().split('\n')[0].strip().upper()
            print(f"[approval] Reading email: '{first_line}'")

            yes_qty_match = re.search(r'YES\s+([A-Z&]+)\s+(\d+)', first_line)
            yes_match = re.search(r'YES\s+([A-Z&]+)', first_line)
            no_match = re.search(r'NO\s+([A-Z&]+)', first_line)

            if yes_qty_match:
                ticker = yes_qty_match.group(1).strip()
                qty = int(yes_qty_match.group(2).strip())
                approvals_with_qty.append((ticker, qty))
                print(f"[approval] APPROVED with QTY: {ticker} x {qty}")
                mail.store(msg_id, '+FLAGS', '\\Seen')
            elif yes_match:
                ticker = yes_match.group(1).strip()
                approvals_need_qty.append(ticker)
                print(f"[approval] APPROVED (asking qty): {ticker}")
                mail.store(msg_id, '+FLAGS', '\\Seen')
            elif no_match:
                ticker = no_match.group(1).strip()
                rejections.append(ticker)
                print(f"[approval] REJECTED: {ticker}")
                mail.store(msg_id, '+FLAGS', '\\Seen')

        mail.logout()

    except Exception as e:
        print(f"[approval] Error checking Gmail: {e}")

    return approvals_with_qty, approvals_need_qty, rejections


def get_stock_details(ticker):
    try:
        ticker_yf = ticker + NSE_SUFFIX
        stock = yf.Ticker(ticker_yf)
        info = stock.info
        current_price = info.get('currentPrice') or info.get('regularMarketPrice')
        if not current_price:
            return None
        return {
            'ticker': ticker,
            'ticker_yf': ticker_yf,
            'name': info.get('longName', ticker),
            'sector': info.get('sector', 'N/A'),
            'current_price': round(float(current_price), 2),
        }
    except Exception as e:
        print(f"[approval] Error fetching details for {ticker}: {e}")
        return None


def send_qty_request_email(ticker, stock_info):
    price = stock_info['current_price']
    subject = f"How many shares of {ticker} do you want? | Rs.{price}/share"

    rows = ''
    for q in [10, 25, 50, 100, 200, 500]:
        rows += f"<tr><td style='padding:8px 14px;border-bottom:1px solid #eee;'>{q} shares</td><td style='padding:8px 14px;border-bottom:1px solid #eee;font-weight:bold;'>Rs.{price*q:,.2f}</td></tr>"

    html = f"""<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"></head>
<body style="font-family:Arial,sans-serif;max-width:600px;margin:auto;background:#f5f5f5;padding:20px;">
  <div style="background:#2E7D32;color:white;padding:24px;border-radius:12px;margin-bottom:20px;">
    <h1 style="margin:0;font-size:20px;">How many shares of {ticker}?</h1>
    <p style="margin:6px 0 0;opacity:0.9;">{stock_info['name']}</p>
  </div>
  <div style="background:white;border-radius:10px;padding:20px;margin-bottom:16px;">
    <h3 style="margin:0 0 12px;">Current Price: Rs.{price:,.2f} per share</h3>
    <table style="width:100%;border-collapse:collapse;">
      <tr style="background:#f8f9fa;">
        <th style="padding:8px 14px;text-align:left;">Quantity</th>
        <th style="padding:8px 14px;text-align:left;">Total Investment</th>
      </tr>
      {rows}
    </table>
  </div>
  <div style="background:#E3F2FD;border:2px dashed #1976D2;border-radius:10px;padding:20px;text-align:center;">
    <strong>Reply to this email with:</strong><br><br>
    <code style="font-size:20px;color:#1976D2;">YES {ticker} [qty]</code><br><br>
    Example: YES {ticker} 50 to buy 50 shares for Rs.{price*50:,.2f}
  </div>
</body>
</html>"""

    send_report_email(subject, html)
    print(f"[approval] Qty request sent for {ticker}")


def get_last_sno(sheet):
    max_sno = 0
    for row in sheet.iter_rows(min_row=3, values_only=True):
        if row[0] and str(row[0]).replace('.', '').isdigit():
            try:
                max_sno = max(max_sno, int(float(row[0])))
            except Exception:
                pass
    return max_sno


def add_stock_to_excel(stock_info, qty):
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
        sheet = wb[HOLDINGS_SHEET]

        buying_price = stock_info['current_price']
        today = datetime.now().strftime('%Y-%m-%d')
        investment_amt = round(buying_price * qty, 2)
        sno = get_last_sno(sheet) + 1

        # Find next row after last S.no entry
        next_row = get_last_sno(sheet) + 2  # +2 because row 1 is header

        # Only write columns A D E F G — B and C left blank for manual Stock Data Type connection
        sheet.cell(row=next_row, column=1).value = sno            # A - S.no
        sheet.cell(row=next_row, column=4).value = buying_price   # D - Buying Price
        sheet.cell(row=next_row, column=5).value = today          # E - Buying Date
        sheet.cell(row=next_row, column=6).value = qty            # F - Qty
        sheet.cell(row=next_row, column=7).value = investment_amt # G - Investment amt
        # B (Industry) and C (Stock) left blank - you connect manually via Excel Stock Data Type
        # H I J K L M are formula columns - auto calculate after you connect stock
        wb.save(EXCEL_FILE_PATH)
        print(f"[approval] Added {stock_info['ticker']} | Rs.{buying_price} x {qty} = Rs.{investment_amt:,.2f}")
        return True

    except Exception as e:
        print(f"[approval] Error adding to Excel: {e}")
        return False


def send_confirmation_email(ticker, stock_info, qty, next_row=None):
    buying_price = stock_info['current_price']
    investment = round(buying_price * qty, 2)
    today = datetime.now().strftime('%d %B %Y')
    if next_row is None:
        next_row = "the new"

    subject = f"{ticker} Added | {qty} shares @ Rs.{buying_price:,.2f} | Total Rs.{investment:,.2f}"

    html = f"""<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"></head>
<body style="font-family:Arial,sans-serif;max-width:600px;margin:auto;background:#f5f5f5;padding:20px;">
  <div style="background:#2E7D32;color:white;padding:24px;border-radius:12px;margin-bottom:20px;">
    <h1 style="margin:0;font-size:20px;">Stock Added to Your Portfolio</h1>
    <p style="margin:6px 0 0;opacity:0.9;">{stock_info['name']}</p>
  </div>
  <div style="background:white;border-radius:10px;padding:20px;margin-bottom:16px;">
    <table style="width:100%;border-collapse:collapse;">
      <tr style="background:#f8f9fa;"><td style="padding:10px 14px;font-weight:bold;width:45%;">Buying Price</td><td style="padding:10px 14px;font-size:18px;font-weight:bold;">Rs.{buying_price:,.2f}</td></tr>
      <tr><td style="padding:10px 14px;font-weight:bold;">Buying Date</td><td style="padding:10px 14px;">{today}</td></tr>
      <tr style="background:#f8f9fa;"><td style="padding:10px 14px;font-weight:bold;">Quantity</td><td style="padding:10px 14px;font-size:18px;font-weight:bold;">{qty} shares</td></tr>
      <tr><td style="padding:10px 14px;font-weight:bold;">Investment Amount</td><td style="padding:10px 14px;font-size:20px;font-weight:bold;color:#1B5E20;">Rs.{investment:,.2f}</td></tr>
    </table>
  </div>

  <div style="background:#FFF8E1;border:2px solid #FFC107;border-radius:10px;padding:20px;margin-bottom:16px;">
    <h2 style="margin:0 0 12px;font-size:16px;">⚠️ One Manual Step Required in Excel</h2>
    <p style="margin:0 0 10px;">The stock has been added at <strong>row {next_row}</strong> in your Excel file with Buying Price, Date, Qty and Investment filled in.</p>
    <p style="margin:0 0 10px;">Now please do this in Excel to connect live data:</p>
    <ol style="margin:0;padding-left:20px;line-height:2;">
      <li>Open your <strong>AI Portfolio .xlsx</strong> file</li>
      <li>Go to <strong>row {next_row}</strong>, column <strong>C (Stock)</strong></li>
      <li>Type <strong>{ticker}</strong> in that cell</li>
      <li>Click the cell, go to <strong>Data</strong> tab → <strong>Stocks</strong> in Stock Types</li>
      <li>Select <strong>NSE</strong> exchange and click the correct stock</li>
      <li>All other columns (Industry, Current Price, Profit, Growth etc.) will auto-populate!</li>
    </ol>
  </div>

  <div style="background:#E8F5E9;padding:14px;border-radius:6px;text-align:center;">
    Once you connect the stock in Excel, it will be fully live in tomorrow's 3:45 PM analysis.
  </div>
</body>
</html>"""

    send_report_email(subject, html)
    print(f"[approval] Confirmation email sent for {ticker}")


def get_last_sno_from_file():
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
        sheet = wb[HOLDINGS_SHEET]
        max_sno = 0
        for row in sheet.iter_rows(min_row=3, max_col=1, values_only=True):
            if row[0] and str(row[0]).replace('.','').isdigit():
                try:
                    max_sno = max(max_sno, int(float(row[0])))
                except Exception:
                    pass
        return max_sno
    except:
        return 0


def update_config_holdings(ticker, stock_info, qty, buying_date):
    """Automatically add newly bought stock to config.py HOLDINGS list"""
    try:
        new_entry = (
            f'    {{"ticker": "{ticker}", '
            f'"name": "{stock_info["name"][:40]}", '
            f'"buying_price": {stock_info["current_price"]}, '
            f'"buying_date": "{buying_date}", '
            f'"qty": {qty}, '
            f'"industry": "{stock_info.get("sector", "N/A")}"}},' + '\n'
        )

        config_content = open('config.py', 'r', encoding='utf-8').read()

        # Insert before the closing ] of HOLDINGS list
        insert_pos = config_content.rfind(']')
        config_content = config_content[:insert_pos] + new_entry + config_content[insert_pos:]

        open('config.py', 'w', encoding='utf-8').write(config_content)
        print(f"[approval] ✅ {ticker} added to config.py HOLDINGS")
    except Exception as e:
        print(f"[approval] Warning: Could not update config.py: {e}")


def process_approvals():
    print("[approval] Checking Gmail for YES/NO replies...")
    approvals_with_qty, approvals_need_qty, rejections = fetch_approval_replies()

    if not approvals_with_qty and not approvals_need_qty and not rejections:
        print("[approval] No new replies found.")
        return

    for ticker, qty in approvals_with_qty:
        stock_info = get_stock_details(ticker)
        if not stock_info:
            print(f"[approval] Could not fetch details for {ticker}. Skipping.")
            continue
        buying_date = datetime.now().strftime('%Y-%m-%d')
        success, sno = add_stock_via_vba(
            ticker=ticker,
            buying_price=stock_info['current_price'],
            buying_date=buying_date,
            qty=qty
        )
        if success:
            from config import USE_GOOGLE_SHEETS
            if USE_GOOGLE_SHEETS:
                from sheets_handler import add_stock_to_holdings
                add_stock_to_holdings(
                    ticker=ticker,
                    stock_name=stock_info.get('name', ticker),
                    industry=stock_info.get('sector', 'N/A'),
                    buying_price=stock_info['current_price'],
                    buying_date=buying_date,
                    qty=qty
                )
            else:
                update_config_holdings(ticker, stock_info, qty, buying_date)
            send_confirmation_email(ticker, stock_info, qty, sno)

    for ticker in approvals_need_qty:
        stock_info = get_stock_details(ticker)
        if not stock_info:
            print(f"[approval] Could not fetch details for {ticker}. Skipping.")
            continue
        send_qty_request_email(ticker, stock_info)

    for ticker in rejections:
        print(f"[approval] {ticker} skipped.")


if __name__ == "__main__":
    process_approvals()